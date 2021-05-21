import json,os
import pandas as pd
from openpyxl import load_workbook

from openpyxl.descriptors import Float
from selenium.webdriver import ActionChains

import tests
from time import sleep
from an.test.ab.HomePage import HomePage
from an.test.ab.ProjectPage import ProjectPage
from an.test.ab.AudiencePage import AudiencePage
from tests.AudienceBuilder import XLUtils

class Aqfer_test(tests.SeleniumTest):
    def test_download_audiences(self):
        project_name = self.configAB["Aqfer"]["project_name"]
        path = os.path.join('testdata', 'sidebar.json')
        print("location:" + path)

        project_page = ProjectPage(self.webdriver, self.config)
        audience_page = AudiencePage(self.webdriver, self.config)
        home_page = HomePage(self.webdriver, self.config)

        with open(path) as file:
            data = json.load(file)
        for item in range(len(data)):
            link = data[item]["link"]
            # Project selection
            new_project_name = project_name + '_' + link + '_QA'
            print(new_project_name)
            sleep(5)
            home_page.searchProjects(new_project_name)
            sleep(10)
            num_aud_panel = len(self.webdriver.find_elements_by_xpath("//div[@class = 'audiences-section']//audience-panel"))
            print(num_aud_panel)
            for i in range (1,num_aud_panel+1):
                audience_name       = self.webdriver.find_element_by_xpath("(//div[@class = 'audience-title ng-binding'])[position()="+str(i)+"]").click()
                # download_link    = self.webdriver.find_element_by_xpath("(//div[@class = 'audiences-section']//div[contains(@class,'download-button')])[position()= "+str(i)+"]")
                # download_link.click()
                # audience_menu_icon = self.webdriver.find_element_by_xpath("//div[text()='" + audience_name + "']//following::div[1]")
                self.webdriver.execute_script("arguments[0].click();", project_page.audience_menu_icon(audience_name))
                self.webdriver.execute_script("arguments[0].click();", project_page.audience_download_icon(audience_name))
                self.webdriver.execute_script("arguments[0].click();", project_page.audience_data(audience_name))






                # project_page.audience_menu_icon(audience_name).click()
                # project_page.audience_download_icon(audience_name).click()
                # project_page.audience_data(audience_name).click()

                # file_path = os.path.abspath(new_project_name + "-" + audience_name + ".xlsx")
                # print(file_path)
                # downloaded_name = new_project_name + " - " + audience_name + ".xlsx"
                # file_path = "C:\\Users\\Abhishek.Gupta\\Documents\\Audience Explorer Document\\OmniQA_Automation\OmniQA\\"
                # print(file_path)
                # complete_file_path_original = file_path + "original\\" + downloaded_name
                # print(complete_file_path_original)
                # file_path_2 = os.path.abspath(downloaded_name)
                # print(file_path_2 + " this is os path_filepath 2" )
                # file_path_3 = self.workspace
                # print(file_path_3 + " this is the self workspace_file_path 3")
                # complete_file_path_clone = file_path + "clone\\" + downloaded_name
                # print(complete_file_path_clone)

                # sleep(10)
                # original_df = pd.read_excel(complete_file_path_original, encoding='latin1')
                # original_df2 = original_df['Individuals']
                # # writer = pd.ExcelWriter('C:\\Users\\Abhishek.Gupta\\Desktop\\count_2.xlsx', engine='xlsxwriter')
                # writer = pd.ExcelWriter(file_path + "result\\result_" + str(i) + ".xlsx" , engine='xlsxwriter')
                # original_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=0, index=False)
                # clone_df = pd.read_excel(complete_file_path_clone,encoding='latin1')
                # clone_df2 = clone_df['Individuals']
                # clone_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=1, index=False)
                # writer.save()

    # def test_download_age(self):
    #     file_path = self.workspace
    #     print(file_path + " this is the self workspace_file_path 3")
    #     Filelists = os.listdir(file_path + "\\original")
    #     print(Filelists)
    #     for i in range(0, len(Filelists)):
    #         print(Filelists[i])
    #         original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=2, index_col=0)
    #         original_target_df = original_df.iloc[:, 2]
    #         book = load_workbook("C:\\Users\\Abhishek.Gupta\\Documents\\Audience Explorer Document\\OmniQA_Automation\\OmniQA\\" + "result\\result_" + str(i + 1) + ".xlsx")
    #         writer = pd.ExcelWriter("C:\\Users\\Abhishek.Gupta\\Documents\\Audience Explorer Document\\OmniQA_Automation\\OmniQA\\" + "result\\result_" + str(i + 1) + ".xlsx", engine='openpyxl')
    #         writer.book = book
    #         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #         original_target_df.to_excel(writer, "Age", startrow=0, startcol=0)
    #
    #         clone_df = pd.read_excel(file_path + "\\clone\\" + Filelists[i], sheet_name=2, index_col=0)
    #         clone_target_df = clone_df.iloc[:, 2]
    #         clone_target_df.to_excel(writer, sheet_name='Age', startrow=0, startcol=6, index=False)
    #
    #
    #         #     writer = pd.ExcelWriter("C:\\Users\\Abhishek.Gupta\\Documents\\Audience Explorer Document\\OmniQA_Automation\\OmniQA\\" + "result\\result_" + str(i+1) + ".xlsx" , engine='openpyxl')
    #         # #     writer = pd.ExcelWriter("C:\\Users\\Abhishek.Gupta\\Documents\\Audience Explorer Document\\OmniQA_Automation\\OmniQA\\" + "result\\result_" + str(i+1) + ".xlsx" , mode='a') as writer
    #         #     dg.to_excel(writer, sheet_name='xyz', startrow=0, startcol=0, index=False)
    #         writer.save()

    def test_download_audience_counts(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace for audience count method")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            complete_file_path_original = file_path + "\\original\\" + Filelists[i]
            print(complete_file_path_original + "This is the complete original file path")
            original_df = pd.read_excel(complete_file_path_original, encoding='latin1')
            original_df2 = original_df['Individuals']
            writer = pd.ExcelWriter(file_path + "\\result\\result_" + str(i+1) + ".xlsx", engine='xlsxwriter')
            original_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=1, index=False)

            complete_file_path_clone = file_path + "\\clone\\" + Filelists[i]
            clone_df = pd.read_excel(complete_file_path_clone,encoding='latin1')
            clone_df2 = clone_df['Individuals']
            clone_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=3, index=False)
            writer.save()


    def test_download_all_sheets_by_target_percentage(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    original_df.sort_values(by=['Target %'], ascending=False, inplace=True)
                    original_target_df = original_df.loc[:, ("Target %", "Index")]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                else:
                    original_df.sort_values(by=['No Filter | Target %'], ascending=False, inplace=True)
                    original_target_df = original_df.loc[:, ("No Filter | Target %", "No Filter | Index")]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                original_target_df.loc['Sum and sorted by Target% of original'] = original_target_df.sum()
                original_target_df.loc['Avg and sorted by Target% of original'] = original_target_df.iloc[:-1, :].mean()

                book = load_workbook(file_path + "\\result\\result_" + str(i + 1) + ".xlsx")
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + str(i + 1) + ".xlsx", engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=1, index=True)

                clone_df = pd.read_excel(file_path +  "\\clone\\" + Filelists[i], sheet_name=j,header=1)
                # clone_target_df = clone_df.iloc[:, 3:5]

                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    clone_df.sort_values(by=['Target %'], ascending=False, inplace=True)

                    clone_target_df = clone_df.loc[:, ("Target %", "Index")]

                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]


                else:
                    clone_df.sort_values(by=['No Filter | Target %'], ascending=False, inplace=True)

                    clone_target_df = clone_df.loc[:, ("No Filter | Target %", "No Filter | Index")]

                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                clone_target_df.loc['Sum and sorted by Target% of clone'] = clone_target_df.sum()
                clone_target_df.loc['Avg and sorted by Target% of clone'] = clone_target_df.iloc[:-1, :].mean()
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=6)

                writer.save()

    def test_download_all_sheets_by_index(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    original_df.sort_values(by=['Index'], ascending=False, inplace=True)
                    original_target_df = original_df.loc[:, ("Target %", "Index")]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                else:
                    original_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)
                    original_target_df = original_df.loc[:, ("No Filter | Target %", "No Filter | Index")]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                original_target_df.loc['Sum and sorted by Index of original'] = original_target_df.sum()
                original_target_df.loc['Avg and sorted by Index of original'] = original_target_df.iloc[:-1, :].mean()

                book = load_workbook(file_path + "\\result\\result_" + str(i + 1) + ".xlsx")
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + str(i + 1) + ".xlsx", engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=11, index=True)

                clone_df = pd.read_excel(file_path +  "\\clone\\" + Filelists[i], sheet_name=j,header=1)
                # clone_target_df = clone_df.iloc[:, 3:5]

                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    clone_df.sort_values(by=['Index'], ascending=False, inplace=True)

                    clone_target_df = clone_df.loc[:, ("Target %", "Index")]

                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]


                else:
                    clone_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)

                    clone_target_df = clone_df.loc[:, ("No Filter | Target %", "No Filter | Index")]

                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                clone_target_df.loc['Sum and sorted by Index of clone'] = clone_target_df.sum()
                clone_target_df.loc['Avg and sorted by Index of clone'] = clone_target_df.iloc[:-1, :].mean()
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=16)

                writer.save()


    def test_download_all_sheets_by_target_percentage_with_other_attributes(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    original_df = original_df.reset_index(drop=True)
                    original_df.sort_values(by=['Target %'], ascending=False, inplace=True)
                    original_df.sort_index(axis=0)
                    original_target_df=original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                else:
                    original_df = original_df.reset_index(drop=True)
                    original_df.sort_values(by=['No Filter | Target %'], ascending=False, inplace=True)
                    original_df.sort_index(axis=0)
                    original_target_df=original_df.iloc[:, 0:5]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                original_target_df.loc['Sum and sorted by Target% of original'] = original_target_df.iloc[:, -2:].sum()
                original_target_df.loc['Avg and sorted by Target% of original'] = original_target_df.iloc[:-1,-2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + str(i + 1) + ".xlsx")
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + str(i + 1) + ".xlsx", engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=1, index=True)

                clone_df = pd.read_excel(file_path +  "\\clone\\" + Filelists[i], sheet_name=j,header=1)
                # clone_target_df = clone_df.iloc[:, 3:5]

                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    clone_df = clone_df.reset_index(drop=True)
                    clone_df.sort_values(by=['Target %'], ascending=False, inplace=True)
                    clone_df.sort_index(axis=0)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]


                else:
                    clone_df = clone_df.reset_index(drop=True)
                    clone_df.sort_values(by=['No Filter | Target %'], ascending=False, inplace=True)
                    clone_df.sort_index(axis=0)
                    clone_target_df = clone_df.iloc[:, 0:5]
                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                clone_target_df.loc['Sum and sorted by Target% of clone'] = clone_target_df.iloc[:, -2:].sum()
                clone_target_df.loc['Avg and sorted by Target% of clone'] = clone_target_df.iloc[:-1, -2:].mean()
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=11)

                writer.save()

    def test_download_all_sheets_by_index_with_other_attributes(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    original_df = original_df.reset_index(drop=True)
                    original_df.sort_values(by=['Index'], ascending=False, inplace=True)
                    original_df.sort_index(axis=0 , inplace=True)
                    original_target_df=original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                else:
                    original_df = original_df.reset_index(drop=True)
                    original_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)
                    original_df.sort_index(axis=0, inplace= True)
                    original_target_df=original_df.iloc[:, 0:5]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                original_target_df.loc['Sum and sorted by Index of original'] = original_target_df.iloc[:, -2:].sum()
                original_target_df.loc['Avg and sorted by Index of original'] = original_target_df.iloc[:-1,-2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + str(i + 1) + ".xlsx")
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + str(i + 1) + ".xlsx", engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=21, index=True)

                clone_df = pd.read_excel(file_path +  "\\clone\\" + Filelists[i], sheet_name=j,header=1)
                # clone_target_df = clone_df.iloc[:, 3:5]

                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    clone_df = clone_df.reset_index(drop=True)
                    clone_df.sort_values(by=['Index'], ascending=False, inplace=True)
                    clone_df.sort_index(axis=0, inplace= True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]


                else:
                    clone_df = clone_df.reset_index(drop=True)
                    clone_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)
                    clone_df.sort_index(axis=0, inplace= True)
                    clone_target_df = clone_df.iloc[:, 0:5]
                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                clone_target_df.loc['Sum and sorted by Index of clone'] = clone_target_df.iloc[:, -2:].sum()
                clone_target_df.loc['Avg and sorted by Index of clone'] = clone_target_df.iloc[:-1, -2:].mean()
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=31)

                writer.save()


