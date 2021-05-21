import openpyxl


# Written By: Ashwini Paunikar
# Date
# Description:Count the number of rows
def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)  # workbook.active
    return sheet.max_row


# Written By: Ashwini Paunikar
# Date
# Description:Count the number of columns
def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)  # workbook.active
    return sheet.max_column


# Written By: Ashwini Paunikar
# Date
# Description:Read the data from xl file
def ReadData(file, sheetname, rownum, colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=colnum).value


# Written By: Ashwini Paunikar
# Date
# Description:Write the data to xl file
def WriteData(file, sheetname,rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=colnum).value = data
    workbook.save(file)
