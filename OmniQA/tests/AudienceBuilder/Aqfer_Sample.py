import json,os
import pandas as pd
from openpyxl import load_workbook

from openpyxl.descriptors import Float
from selenium.webdriver import ActionChains

import tests
from time import sleep
from an.test.ab.HomePage import HomePage
from an.test.ab.ProjectPage import ProjectPage
from an.test.ab.AudiencePage import AudiencePage
from tests.AudienceBuilder import XLUtils

class Aqfer_test(tests.SeleniumTest):
    # def test_a_download_audience_counts_org(self):
    #     file_path = self.workspace
    #     print(file_path + " this is the self workspace for audience count method")
    #     Filelists = os.listdir(file_path + "\\original")
    #     print(Filelists)
    #     for i in range(0, len(Filelists)):
    #         print(Filelists[i])
    #         complete_file_path_original = file_path + "\\original\\" + Filelists[i]
    #         print(complete_file_path_original + "This is the complete original file path")
    #         original_df = pd.read_excel(complete_file_path_original, encoding='latin1')
    #         original_df2 = original_df['Individuals']
    #         writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i] , engine='xlsxwriter')
    #         original_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=1, index=False)
    #         writer.save()
    #
    # def test_b_download_audience_counts_clone(self):
    #     file_path = self.workspace
    #     print(file_path + " this is the self workspace for audience count method")
    #     Filelists = os.listdir(file_path + "\\original")
    #     Filelists_cl  = os.listdir(file_path + "\\clone")
    #     print(Filelists_cl)
    #     for i in range(0, len(Filelists_cl)):
    #         print(Filelists_cl[i])
    #         complete_file_path_clone = file_path + "\\clone\\" + Filelists_cl[i]
    #         print(complete_file_path_clone + "This is the complete original file path")
    #         clone_df = pd.read_excel(complete_file_path_clone, encoding='latin1')
    #         clone_df2 = clone_df['Individuals']
    #         book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
    #         writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i] , engine='openpyxl')
    #         writer.book = book
    #         writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #         clone_df2.to_excel(writer, sheet_name='Individual_Count', startrow=0, startcol=3, index=False)
    #         writer.save()

    def test_c_download_all_sheets_by_target_percentage_with_other_attributes_org(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    original_df = original_df.reset_index(drop=True)
                    # original_df.sort_values(by=['Target %', 'Index'], ascending=False, inplace=True)
                    # original_df.sort_index(axis=0)
                    original_target_df=original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                else:
                    original_df = original_df.reset_index(drop=True)
                    # original_df.sort_values(by=['No Filter | Target %', 'No Filter | Index'], ascending=False, inplace=True)
                    # original_df.sort_index(axis=0)
                    original_target_df=original_df.iloc[:, 0:5]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]

                original_target_df.loc['Sum and sorted by Target% of original'] = original_target_df.iloc[:, -2:].sum()
                original_target_df.loc['Avg and sorted by Target% of original'] = original_target_df.iloc[:-1,-2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=17, startcol=1, index=True)
                writer.save()

    def test_d_download_all_sheets_by_target_percentage_with_other_attributes_clone(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        Filelists_cl = os.listdir(file_path + "\\clone")
        print(Filelists_cl)
        for i in range(0, len(Filelists_cl)):
            print(Filelists_cl[i])
            xl = pd.ExcelFile(file_path + "\\clone\\" + Filelists_cl[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                clone_df = pd.read_excel(file_path + "\\clone\\" + Filelists_cl[i], sheet_name=j,header=1)
                # original_target_df = original_df.iloc[:, 3:5]
                if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
                    clone_df = clone_df.reset_index(drop=True)
                    # clone_df.sort_values(by=['Target %','Index'], ascending=False, inplace=True)
                    # clone_df.sort_index(axis=0)
                    clone_target_df=clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                else:
                    clone_df = clone_df.reset_index(drop=True)
                    # clone_df.sort_values(by=['No Filter | Target %','No Filter | Index'], ascending=False, inplace=True)
                    # clone_df.sort_index(axis=0)
                    clone_target_df =clone_df.iloc[:, 0:5]
                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]

                clone_target_df.loc['Sum and sorted by Target% of clone'] = clone_target_df.iloc[:, -2:].sum()
                clone_target_df.loc['Avg and sorted by Target% of clone'] = clone_target_df.iloc[:-1,-2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=17, startcol=11, index=True)
                writer.save()

    # def test_e_download_all_sheets_by_index_with_other_attributes_org(self):
    #     file_path = self.workspace
    #     print(file_path + " this is the self workspace_file_path 3")
    #     Filelists = os.listdir(file_path + "\\original")
    #     print(Filelists)
    #     for i in range(0, len(Filelists)):
    #         print(Filelists[i])
    #         xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
    #         print(xl.sheet_names)
    #         print(len(xl.sheet_names))
    #         for j in range(1, len(xl.sheet_names)):
    #             original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
    #             # original_target_df = original_df.iloc[:, 3:5]
    #             if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
    #                 original_df = original_df.reset_index(drop=True)
    #                 # original_df.sort_values(by=['Index','Target %'], ascending=False, inplace=True)
    #                 original_df.sort_values(by=['Index'], ascending=False, inplace=True)
    #                 original_df.sort_index(axis=0 , inplace=True)
    #                 original_target_df=original_df.iloc[:, 0:8]
    #                 original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
    #                 original_target_df = original_target_df.iloc[0:10, :]
    #
    #             else:
    #                 original_df = original_df.reset_index(drop=True)
    #                 # original_df.sort_values(by=['No Filter | Index', 'No Filter | Target %'], ascending=False, inplace=True)
    #                 original_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)
    #                 original_df.sort_index(axis=0, inplace= True)
    #                 original_target_df=original_df.iloc[:, 0:5]
    #                 original_target_df['No Filter | Target %'] = original_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
    #                 original_target_df = original_target_df.iloc[0:10, :]
    #
    #             original_target_df.loc['Sum and sorted by Index of original'] = original_target_df.iloc[:, -2:].sum()
    #             original_target_df.loc['Avg and sorted by Index of original'] = original_target_df.iloc[:-1,-2:].mean()
    #
    #             book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
    #             writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
    #             writer.book = book
    #             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #             original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=21, index=True)
    #             writer.save()
    #
    #
    # def test_f_download_all_sheets_by_index_with_other_attributes_clone(self):
    #     file_path = self.workspace
    #     print(file_path + " this is the self workspace_file_path 3")
    #     Filelists = os.listdir(file_path + "\\original")
    #     Filelists_cl = os.listdir(file_path + "\\clone")
    #     print(Filelists_cl)
    #     for i in range(0, len(Filelists_cl)):
    #         print(Filelists_cl[i])
    #         xl = pd.ExcelFile(file_path + "\\clone\\" + Filelists_cl[i])
    #         print(xl.sheet_names)
    #         print(len(xl.sheet_names))
    #         for j in range(1, len(xl.sheet_names)):
    #             clone_df = pd.read_excel(file_path + "\\clone\\" + Filelists_cl[i], sheet_name=j,header=1)
    #             # original_target_df = original_df.iloc[:, 3:5]
    #             if (xl.sheet_names[j] == 'Advanced Audience Data' or xl.sheet_names[j] == 'Purchase Behavior' or xl.sheet_names[j] == 'Location Data' or xl.sheet_names[j] == 'TV - Series' or xl.sheet_names[j] == 'TV - Network' or xl.sheet_names[j] == 'TV - Category'):
    #                 clone_df = clone_df.reset_index(drop=True)
    #                 # clone_df.sort_values(by=['Index', 'Target %'], ascending=False, inplace=True)
    #                 clone_df.sort_values(by=['Index'], ascending=False, inplace=True)
    #                 # clone_df.sort_index(axis=0 , inplace=True)
    #                 clone_target_df=clone_df.iloc[:, 0:8]
    #                 clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
    #                 clone_target_df = clone_target_df.iloc[0:10, :]
    #
    #             else:
    #                 clone_df = clone_df.reset_index(drop=True)
    #                 # clone_df.sort_values(by=['No Filter | Index', 'No Filter | Target %'], ascending=False, inplace=True)
    #                 clone_df.sort_values(by=['No Filter | Index'], ascending=False, inplace=True)
    #                 clone_df.sort_index(axis=0, inplace= True)
    #                 clone_target_df=clone_df.iloc[:, 0:5]
    #                 clone_target_df['No Filter | Target %'] = clone_target_df.loc[:,("No Filter | Target %")].apply(lambda x: x * 100)
    #                 clone_target_df = clone_target_df.iloc[0:10, :]
    #
    #             clone_target_df.loc['Sum and sorted by Index of clone'] = clone_target_df.iloc[:, -2:].sum()
    #             clone_target_df.loc['Avg and sorted by Index of clone'] = clone_target_df.iloc[:-1,-2:].mean()
    #
    #             book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
    #             writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
    #             writer.book = book
    #             writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #             clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=31, index=True)
    #             writer.save()