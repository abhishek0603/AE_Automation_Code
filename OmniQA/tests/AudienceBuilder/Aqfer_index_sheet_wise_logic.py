import json,os
import pandas as pd
from openpyxl import load_workbook

from openpyxl.descriptors import Float
from selenium.webdriver import ActionChains

import tests
from time import sleep
from an.test.ab.HomePage import HomePage
from an.test.ab.ProjectPage import ProjectPage
from an.test.ab.AudiencePage import AudiencePage
from tests.AudienceBuilder import XLUtils
class Aqfer_test(tests.SeleniumTest):
    def test_e_download_all_sheets_by_index_with_other_attributes_org(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        print(Filelists)
        for i in range(0, len(Filelists)):
            print(Filelists[i])
            xl = pd.ExcelFile(file_path + "\\original\\" + Filelists[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                original_df = pd.read_excel(file_path + "\\original\\" + Filelists[i], sheet_name=j,header=1)
                if (xl.sheet_names[j] == 'Advanced Audience Data'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Index', 'Category', 'Subcategory'], ascending=[False, True, True],inplace=True)
                    original_target_df = original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Category'] + original_target_df['Subcategory'] + original_target_df['Attribute Description']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)

                elif (xl.sheet_names[j] == 'Purchase Behavior'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Target %', 'Category', 'Brand', 'Index'], ascending=[False, True, True, False],inplace=True)
                    original_target_df = original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Category'] + original_target_df['Subcategory'] + original_target_df['Product Name']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)

                elif(xl.sheet_names[j] == 'Location Data'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Index', 'Category', 'Brand'], ascending=[False, True, True],inplace=True)
                    original_target_df = original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Category'] + original_target_df['Brand']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)

                elif (xl.sheet_names[j] == 'TV - Series'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Index', 'Index Type', 'Network', 'Category', 'Subcategory'],ascending=[False, True, True, True, True], inplace=True)
                    original_target_df = original_df.iloc[:, 0:9]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Index Type'] + original_target_df['Network'] + original_target_df['Series'] + original_target_df['Program Type'] + original_target_df['Category'] + original_target_df['Subcategory']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)

                elif (xl.sheet_names[j] == 'TV - Network'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Index', 'Index Type', 'Network'], ascending=[False, True, True],inplace=True)
                    original_target_df = original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Index Type'] + original_target_df['Network']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)

                elif (xl.sheet_names[j] == 'TV - Category'):
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['Index', 'Index Type', 'Category', 'Subcategory'],ascending=[False, True, True, True], inplace=True)
                    original_target_df = original_df.iloc[:, 0:8]
                    original_target_df['Target %'] = original_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Index Type'] + original_target_df['Category'] + original_target_df['Subcategory']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)
                else:
                    original_df = original_df.reset_index()
                    # original_df.sort_values(by=['No Filter | Index', 'Att Type', 'Category', 'Attribute Value'],ascending=[False, True, True, True], inplace=True)
                    original_target_df = original_df.iloc[:, 0:6]
                    original_target_df['No Filter | Target %'] = original_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    original_target_df = original_target_df.iloc[0:10, :]
                    original_target_df['Concate'] = original_target_df['Att Type'] + original_target_df['Category'] + original_target_df['Attribute Value']
                    # rank = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
                    original_target_df['Rank'] = range(1,len(original_target_df)+1)


                # original_target_df.loc['Sum and sorted by Index of original'] = original_target_df.iloc[:, -2:].sum()
                original_target_df.loc[len(original_target_df) +2, 'index'] = "Sorted by Index of original"
                # original_target_df.loc['Avg and sorted by Index of original'] = original_target_df.iloc[:-1, -2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                original_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=27, index=False)
                writer.save()

    def test_f_download_all_sheets_by_index_with_other_attributes_clone(self):
        file_path = self.workspace
        print(file_path + " this is the self workspace_file_path 3")
        Filelists = os.listdir(file_path + "\\original")
        Filelists_cl = os.listdir(file_path + "\\clone")
        print(Filelists_cl)
        for i in range(0, len(Filelists_cl)):
            print(Filelists_cl[i])
            xl = pd.ExcelFile(file_path + "\\clone\\" + Filelists_cl[i])
            print(xl.sheet_names)
            print(len(xl.sheet_names))
            for j in range(1, len(xl.sheet_names)):
                clone_df = pd.read_excel(file_path + "\\clone\\" + Filelists_cl[i], sheet_name=j,header=1)
                if (xl.sheet_names[j] == 'Advanced Audience Data'):
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['Index', 'Category', 'Subcategory'], ascending=[False, True, True],inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] = clone_target_df['Category'] + clone_target_df['Subcategory'] + clone_target_df['Attribute Description']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                elif (xl.sheet_names[j] == 'Purchase Behavior'):
                    clone_df = clone_df.reset_index()
                    # original_df.sort_values(by=['Target %', 'Category', 'Brand', 'Index'], ascending=[False, True, True, False],inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] = clone_target_df['Category'] + clone_target_df['Subcategory'] + clone_target_df['Product Name']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                elif(xl.sheet_names[j] == 'Location Data'):
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['Index', 'Category', 'Brand'], ascending=[False, True, True],inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] = clone_target_df['Category'] + clone_target_df['Brand']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                elif (xl.sheet_names[j] == 'TV - Series'):
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['Index', 'Index Type', 'Network', 'Category', 'Subcategory'],ascending=[False, True, True, True, True], inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:9]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] = clone_target_df['Index Type'] + clone_target_df['Network'] + clone_target_df['Series'] + clone_target_df['Program Type'] + clone_target_df['Category'] + clone_target_df['Subcategory']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                elif (xl.sheet_names[j] == 'TV - Network'):
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['Index', 'Index Type', 'Network'], ascending=[False, True, True],inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] = clone_target_df['Index Type'] + clone_target_df['Network']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                elif (xl.sheet_names[j] == 'TV - Category'):
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['Index', 'Index Type', 'Category', 'Subcategory'],ascending=[False, True, True, True], inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:8]
                    clone_target_df['Target %'] = clone_target_df.loc[:, ("Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] =  clone_target_df['Index Type'] + clone_target_df['Category'] + clone_target_df['Subcategory']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""

                else:
                    clone_df = clone_df.reset_index()
                    # clone_df.sort_values(by=['No Filter | Index', 'Att Type', 'Category', 'Attribute Value'],ascending=[False, True, True, True], inplace=True)
                    clone_target_df = clone_df.iloc[:, 0:6]
                    clone_target_df['No Filter | Target %'] = clone_target_df.loc[:, ("No Filter | Target %")].apply(lambda x: x * 100)
                    clone_target_df = clone_target_df.iloc[0:10, :]
                    clone_target_df['Concate'] =  clone_target_df['Att Type'] + clone_target_df['Category'] + clone_target_df['Attribute Value']
                    clone_target_df['Rank'] = ""
                    clone_target_df['Diff'] = ""



                # clone_target_df.loc['Sum and sorted by Index of original'] = clone_target_df.iloc[:, -2:].sum()
                clone_target_df.loc[len(clone_target_df) +2, 'index'] = "Sorted by Index of clone"
                # clone_target_df.loc['Avg and sorted by Index of original'] = clone_target_df.iloc[:-1, -2:].mean()

                book = load_workbook(file_path + "\\result\\result_" + Filelists[i])
                writer = pd.ExcelWriter(file_path + "\\result\\result_" + Filelists[i], engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                clone_target_df.to_excel(writer, xl.sheet_names[j], startrow=0, startcol=40, index=False)
                writer.save()